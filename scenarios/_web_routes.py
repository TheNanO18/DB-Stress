"""
Locust 웹 UI 확장 — 모니터링 대시보드 라우트 및 배너 주입

import 시점에 @events.init 데코레이터가 자동으로 리스너를 등록합니다.
"""

import io
import json
import os
from datetime import datetime

from locust import events

from core.metrics_store import get_metrics_store

_metrics = get_metrics_store()


@events.init.add_listener
def on_init(environment, **kwargs):
    """Locust 웹 서버에 모니터링 대시보드 라우트를 추가합니다."""
    if environment.web_ui is None:
        return

    from flask import Response

    _template_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "templates"
    )

    @environment.web_ui.app.route("/monitor")
    def monitor_page():
        html_path = os.path.join(_template_dir, "monitor.html")
        with open(html_path, "r", encoding="utf-8") as f:
            return Response(f.read(), content_type="text/html; charset=utf-8")

    @environment.web_ui.app.route("/monitor/data")
    def monitor_data():
        data = _metrics.snapshot()
        data["mode"] = _metrics.mode
        return Response(
            json.dumps(data), content_type="application/json"
        )

    @environment.web_ui.app.route("/monitor/export/excel")
    def monitor_export_excel():
        """모니터링 시계열 데이터를 Excel(.xlsx)로 내보냅니다."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        data = _metrics.snapshot()
        timestamps = data.pop("timestamps", [])
        data.pop("mode", None)
        container_labels = data.pop("container_labels", [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Monitor Data"

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")

        # 컬럼 구성: Time + 각 metric 키
        metric_keys = [k for k in data.keys() if isinstance(data[k], list)]
        headers = ["Time"] + metric_keys

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 데이터 행
        for row_idx, ts in enumerate(timestamps, 2):
            time_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row_idx, column=1, value=time_str)
            for col_idx, key in enumerate(metric_keys, 2):
                values = data[key]
                val = values[row_idx - 2] if (row_idx - 2) < len(values) else None
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 컬럼 너비 자동 조정
        ws.column_dimensions["A"].width = 22
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cubrid_monitor_{now_str}.xlsx"
        return Response(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @environment.web_ui.app.route("/monitor/export/csv")
    def monitor_export_csv():
        """모니터링 시계열 데이터를 CSV로 내보냅니다."""
        data = _metrics.snapshot()
        timestamps = data.pop("timestamps", [])
        data.pop("mode", None)
        data.pop("container_labels", [])

        metric_keys = [k for k in data.keys() if isinstance(data[k], list)]
        lines = [",".join(["Time"] + metric_keys)]

        for i, ts in enumerate(timestamps):
            time_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
            row = [time_str]
            for key in metric_keys:
                values = data[key]
                row.append(str(values[i]) if i < len(values) else "")
            lines.append(",".join(row))

        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cubrid_monitor_{now_str}.csv"
        return Response(
            "\n".join(lines),
            content_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # Locust 메인 페이지에 Monitor 링크 배너를 주입하는 미들웨어
    _original_app = environment.web_ui.app

    # 시나리오 클래스명 → 한국어 설명 매핑
    _class_descriptions = {
        "BulkInsertUser": "대량 INSERT — 디스크 I/O, 인덱스 부하",
        "ReadIntensiveUser": "읽기 집중 — PK SELECT로 최대 TPS 측정",
        "LockContentionUser": "Lock 경합 — 동일 행(1~3) 동시 UPDATE (2명 이상, 단독 실행 권장)",
        "HeavyQueryUser": "Heavy 쿼리 — 셀프 조인, 풀스캔, 대량 정렬",
        "ConnectionChurnUser": "Connection Churn — 연결 생성/해제 반복",
        "CrudMixUser": "CRUD 종합 — INSERT/SELECT/UPDATE/DELETE 균등 실행",
        "CreateOnlyUser": "Create 단독 — INSERT만 반복 실행",
        "ReadOnlyUser": "Read 단독 — PK SELECT만 반복 실행",
        "UpdateOnlyUser": "Update 단독 — UPDATE만 반복 실행",
        "DeleteOnlyUser": "Delete 단독 — DELETE만 반복 실행",
        "DBMonitorUser": "DB 모니터 — 부하 없이 상태만 관찰",
    }
    _desc_json = json.dumps(_class_descriptions, ensure_ascii=False)

    @_original_app.after_request
    def inject_monitor_link(response):
        if response.content_type and "text/html" in response.content_type:
            inject_html = (
                '<div id="monitor-banner" style="'
                'position:fixed;top:0;left:0;right:0;z-index:9999;'
                'background:#16213e;border-bottom:2px solid #00d4aa;'
                'padding:6px 16px;text-align:center;font-size:13px;'
                'font-family:Segoe UI,sans-serif;">'
                '<a href="/monitor" style="color:#00d4aa;text-decoration:none;font-weight:bold;">'
                '&#128202; CUBRID Stress Monitor Dashboard &rarr;'
                '</a></div>'
                '<script>document.body.style.paddingTop="36px";</script>'
                '<script>'
                '(function(){'
                '  var desc=' + _desc_json + ';'
                '  function addDesc(){'
                '    var cells=document.querySelectorAll("td.MuiTableCell-root");'
                '    cells.forEach(function(td){'
                '      var txt=td.textContent.trim();'
                '      if(desc[txt] && !td.dataset.descAdded){'
                '        td.dataset.descAdded="1";'
                '        var span=document.createElement("span");'
                '        span.style.cssText='
                '          "color:#888;font-size:0.85em;margin-left:8px;'
                '           font-style:italic;";'
                '        span.textContent="("+desc[txt]+")";'
                '        td.appendChild(span);'
                '      }'
                '    });'
                '  }'
                '  var observer=new MutationObserver(function(){addDesc();});'
                '  observer.observe(document.body,{childList:true,subtree:true});'
                '  addDesc();'
                '})();'
                '</script>'
                '<script>'
                '(function(){'
                '  var LOCK="LockContentionUser",MON="DBMonitorUser";'
                '  var busy=false;'
                ''
                '  function getRows(){'
                '    var map={};'
                '    document.querySelectorAll("td.MuiTableCell-root").forEach(function(td){'
                '      var raw=td.textContent.trim();'
                '      var name=raw.split("(")[0].trim();'
                '      if(!name.endsWith("User"))return;'
                '      var tr=td.closest("tr");'
                '      if(!tr)return;'
                '      var cb=tr.querySelector("input[type=checkbox]");'
                '      if(!cb)return;'
                '      map[name]=cb;'
                '    });'
                '    return map;'
                '  }'
                ''
                '  function setInput(el,val){'
                '    var setter=Object.getOwnPropertyDescriptor('
                '      window.HTMLInputElement.prototype,"value").set;'
                '    setter.call(el,val);'
                '    el.dispatchEvent(new Event("input",{bubbles:true}));'
                '  }'
                ''
                '  function init(){'
                '    var rows=getRows();'
                '    if(Object.keys(rows).length<2)return;'
                '    var anyNew=false;'
                '    Object.keys(rows).forEach(function(n){'
                '      if(!rows[n].dataset.lockInited) anyNew=true;'
                '    });'
                '    if(!anyNew)return;'
                ''
                '    busy=true;'
                '    Object.keys(rows).forEach(function(n){'
                '      if(rows[n].checked) rows[n].click();'
                '      rows[n].dataset.lockInited="1";'
                '    });'
                '    setTimeout(function(){busy=false;},200);'
                ''
                '    Object.keys(rows).forEach(function(n){'
                '      rows[n].addEventListener("click",function(){'
                '        if(busy)return;'
                '        setTimeout(function(){'
                '          busy=true;'
                '          var r=getRows();'
                '          if(n===LOCK && r[LOCK] && r[LOCK].checked){'
                '            Object.keys(r).forEach(function(k){'
                '              if(k!==LOCK && k!==MON && r[k].checked) r[k].click();'
                '            });'
                '          } else if(n!==LOCK && n!==MON && r[n] && r[n].checked){'
                '            if(r[LOCK] && r[LOCK].checked) r[LOCK].click();'
                '          }'
                '          busy=false;'
                '        },50);'
                '      });'
                '    });'
                '  }'
                ''
                '  function watchStartForm(){'
                '    var dlg=document.querySelector("div.MuiDialog-root");'
                '    if(!dlg || dlg.dataset.lockPatched)return;'
                '    dlg.dataset.lockPatched="1";'
                '    var rows=getRows();'
                '    if(!rows[LOCK] || !rows[LOCK].checked)return;'
                '    var inputs=dlg.querySelectorAll("input[type=number],input.MuiInputBase-input");'
                '    if(inputs.length>=1) setInput(inputs[0],"3");'
                '    if(inputs.length>=2) setInput(inputs[1],"3");'
                '  }'
                ''
                '  var obs=new MutationObserver(function(){init();watchStartForm();});'
                '  obs.observe(document.body,{childList:true,subtree:true});'
                '  init();'
                '})();'
                '</script>'
            )
            data = response.get_data(as_text=True)
            if '/monitor' not in data and '<body' in data.lower():
                data = data.replace('</body>', inject_html + '</body>')
                response.set_data(data)
        return response

    print("[Init] 모니터링 대시보드: http://localhost:8089/monitor")
